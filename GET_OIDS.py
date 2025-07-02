from pysnmp.hlapi import *
import sys
from datetime import datetime
from openpyxl import load_workbook
import csv


now = datetime.now()
time2 = now.strftime("%d-%m-%Y %H:%M")
def walk(oid, ip, host):
    data=[]
    data.append(host)
    data.append(ip)
    try:
        iterator = getCmd(SnmpEngine(),
                          UsmUserData('user', '1234', '1234'),
                          UdpTransportTarget((ip, 161)),
                          ContextData(),
                          ObjectType(ObjectIdentity(oid)),
                          lookupMib=False,
                          lexicographicMode=False)
        errorIndication, errorStatus, errorIndex, varBinds = iterator
        if errorIndication:
            #print(errorIndication)
            data.append(errorIndication)
        elif errorStatus:
            #print('%s at %s' % (errorStatus.prettyPrint(),errorIndex and varBinds[int(errorIndex) - 1][0] or '?'))
            data.append('%s at %s' % (errorStatus.prettyPrint(),
                                errorIndex and varBinds[int(errorIndex) - 1][0] or '?'))
        else:
            for varBind in varBinds:
                #print(' = '.join([x.prettyPrint() for x in varBind]))
                data.append(' = '.join([x.prettyPrint() for x in varBind]).split(";")[3].replace('+0',''))
        data.append(time2)
        return data
    except Exception as e:
        print(e)

oid='1.3.6.1.4.1'

file_csv=open('test.csv','w',encoding='utf-8',newline='')
write_csv=csv.writer(file_csv)


try:
    wb = load_workbook('LIST.xlsx')
    wb.active = 0
    ws = wb.active
    # print(wb.active)

    data_list = []
    'Pasamos la hoja al dictionario'
    headers = [cell.value for cell in ws[1]]
    # print(headers)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dit = {}
        for header, value in zip(headers, row):
            row_dit[header] = value
            data_list.append(row_dit)
    data_optz = []
    data_list_optz = []
    for s in data_list:
        data_optz = []
        for k, v in s.items():
            if k == 'Hostname':
                data_optz.append(v)
            if k == 'Nombre Equipo':
                data_optz.append(v)
            if k == 'IP':
                data_optz.append(v)
        data_list_optz.append(data_optz)
    #print(data_list_optz)
    data_limpio = []
    for i in data_list_optz:
        if i not in data_limpio:
            data_limpio.append(i)

    #print("LECTURA EXCEL XLSX: ",data_limpio)
    #logging.debug("LECTURA EXCEL XLSX: {0} ".format(data_limpio))
    wb.close()

    for x in data_limpio:
        print(walk(oid, x[2], x[1]))
        write_csv.writerow(walk(oid, x[2], x[1]))
        status = []


except Exception as e:
    print(e)
    #messagebox.showerror("Error", "Error al leer excel "f"{e}")
    #logging.debug("Error al leer el archivo Excel: %s", str(e))

























